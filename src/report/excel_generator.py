"""
Excel Generator: Phục vụ xuất báo cáo trực tiếp vào file template Form.xlsx
"""
import os
import shutil
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def fill_fpt_template(
    survey_df: pd.DataFrame,
    dssv_df: pd.DataFrame = None,
    template_path: str = "Form.xlsx",
    output_path: str = "output_report.xlsx",
    ky: str = "Spring 2026",
    coso: str = "FPTU CẦN THƠ",
    ai_group_result: dict = None
) -> str:
    """
    Đọc file Form.xlsx, điền dữ liệu khảo sát và trả về đường dẫn file kết quả.
    """
    # 1. Sao chép template sang file output
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Không tìm thấy file template: {template_path}")
    
    shutil.copy(template_path, output_path)
    
    # 2. Mở workbook bằng openpyxl
    wb = openpyxl.load_workbook(output_path)
    
    # 3. Phân loại ngành và tính toán thống kê
    df_cntt = survey_df[survey_df["Major"] == "Công nghệ thông tin"]
    df_biz  = survey_df[survey_df["Major"] == "Quản trị kinh doanh"]
    df_lang = survey_df[survey_df["Major"] == "Ngôn ngữ"]
    
    # Đếm số lượng SV từ DSSV (nếu có)
    sv_cntt = 0
    sv_biz  = 0
    sv_lang = 0
    if dssv_df is not None and len(dssv_df) > 0:
        col_nganh = None
        for col in dssv_df.columns:
            col_l = col.lower()
            if "ngành" in col_l or "nganh" in col_l or "major" in col_l:
                col_nganh = col
                break
        if col_nganh:
            dssv_df = dssv_df.copy()
            def _normalize_major(x):
                val = str(x).strip().upper()
                if val.startswith("BIT") or any(k in val for k in ["CNTT", "IT", "SE", "GD", "IA", "AI", "KTPM", "DESIGN"]):
                    return "Công nghệ thông tin"
                if val.startswith("BBA") or any(k in val for k in ["QTKD", "BIZ", "BUS", "MKT", "IB", "MC", "HA", "HM", "TÀI CHÍNH", "MARKETING"]):
                    return "Quản trị kinh doanh"
                if val.startswith("BEN") or any(k in val for k in ["NNA", "NNH", "NNK", "NNT", "ENGLISH", "LANGUAGE", "NGÔN NGỮ"]):
                    return "Ngôn ngữ"
                
                val_lower = val.lower()
                if any(k in val_lower for k in ["it", "công nghệ", "cntt", "information", "phần mềm", "đồ họa"]):
                    return "Công nghệ thông tin"
                if any(k in val_lower for k in ["business", "kinh tế", "quản trị", "qtkd", "marketing", "kinh doanh"]):
                    return "Quản trị kinh doanh"
                if any(k in val_lower for k in ["ngôn ngữ", "language", "nna", "nnh", "nnt", "tiếng anh", "tiếng nhật", "tiếng hàn"]):
                    return "Ngôn ngữ"
                return "Khác"

            dssv_df["_Major_Norm"] = dssv_df[col_nganh].apply(_normalize_major)
            sv_cntt = (dssv_df["_Major_Norm"] == "Công nghệ thông tin").sum()
            sv_biz  = (dssv_df["_Major_Norm"] == "Quản trị kinh doanh").sum()
            sv_lang = (dssv_df["_Major_Norm"] == "Ngôn ngữ").sum()
    
    # 4. Hàm hỗ trợ tính toán metrics cho 8 câu hỏi
    def calculate_question_metrics(df_sub):
        metrics = []
        for q in range(1, 9):
            score_col = f"Q{q}_Score"
            if score_col in df_sub.columns:
                series = df_sub[score_col].dropna()
                n_valid = len(series)
                counts = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0, 0: 0}
                for v in series:
                    if v in counts:
                        counts[v] += 1
                
                # Không có ý kiến (rating trống hoặc 0)
                counts[0] = len(df_sub) - n_valid
                
                avg_score = series.mean() if n_valid > 0 else 0.0
                sat_rate = (counts[5] + counts[4]) / n_valid if n_valid > 0 else 0.0
                metrics.append({
                    "counts": counts,
                    "avg": avg_score,
                    "rate": sat_rate,
                    "valid": n_valid
                })
            else:
                metrics.append({
                    "counts": {5: 0, 4: 0, 3: 0, 2: 0, 1: 0, 0: len(df_sub)},
                    "avg": 0.0,
                    "rate": 0.0,
                    "valid": 0
                })
        return metrics

    metrics_all  = calculate_question_metrics(survey_df)
    metrics_cntt = calculate_question_metrics(df_cntt)
    metrics_biz  = calculate_question_metrics(df_biz)
    metrics_lang = calculate_question_metrics(df_lang)

    # 5. Hàm ghi dữ liệu rating vào sheet mẫu
    def fill_rating_table(ws, metrics, start_row=8):
        for idx, m in enumerate(metrics):
            r = start_row + idx
            c = m["counts"]
            
            # Ghi số lượng bình chọn vào cột F, G, H, I, J, K (tương ứng với 5, 4, 3, 2, 1, 0)
            ws.cell(row=r, column=6, value=c[5])
            ws.cell(row=r, column=7, value=c[4])
            ws.cell(row=r, column=8, value=c[3])
            ws.cell(row=r, column=9, value=c[2])
            ws.cell(row=r, column=10, value=c[1])
            ws.cell(row=r, column=11, value=c[0])
            
            # Điểm TB (Col L / column 12)
            cell_avg = ws.cell(row=r, column=12, value=round(m["avg"], 2))
            cell_avg.number_format = '0.00'
            
            # Tỷ lệ HL (Col M / column 13)
            cell_rate = ws.cell(row=r, column=13, value=m["rate"])
            cell_rate.number_format = '0.00%'

    # 6. Cập nhật Metadata (Kỳ học, Cơ sở) trên các sheet
    for sname in wb.sheetnames:
        ws = wb[sname]
        # Tìm và thay thế các ô chứa text mẫu
        for r in range(1, 6):
            for c in range(1, 15):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str):
                    if "<<FPTU CT/ FGW CT>>" in val or "ĐẠI HỌC <<FPT" in val or "<<FPTU/ FGW>>" in val:
                        ws.cell(row=r, column=c, value=val.replace("<<FPTU CT/ FGW CT>>", coso).replace("<<FPTU/ FGW>>", coso))
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, str) and ("<<>>" in val or "Kỳ học: <<>>" in val):
                        ws.cell(row=r, column=c, value=val.replace("<<>>", ky))
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, str) and "…../…../ 2026" in val:
                        import datetime
                        today_str = datetime.date.today().strftime("%d/%m/%Y")
                        ws.cell(row=r, column=c, value=val.replace("…../…../ 2026", today_str))

    # 7. Ghi dữ liệu cho sheet 'TONG HOP'
    ws_total = wb['TONG HOP']
    
    # Cập nhật số lượng SV phản hồi ở Row 4
    val_r4 = ws_total.cell(row=4, column=4).value
    if isinstance(val_r4, str) and "<<>>" in val_r4:
        ws_total.cell(row=4, column=4, value=val_r4.replace("<<>>", str(len(survey_df))))
    elif isinstance(val_r4, str) and "sinh viên phản hồi" in val_r4:
        # Nếu đã bị replace phần metadata, ta có thể viết đè lại
        ws_total.cell(row=4, column=4, value=f"TỔNG HỢP CHUNG ({len(survey_df)} sinh viên phản hồi)")
    
    # Section I: Tỷ lệ phản hồi
    if sv_cntt > 0:
        ws_total.cell(row=7, column=8, value=sv_cntt)
    ws_total.cell(row=7, column=9, value=len(df_cntt))
    
    if sv_biz > 0:
        ws_total.cell(row=8, column=8, value=sv_biz)
    ws_total.cell(row=8, column=9, value=len(df_biz))
    
    if sv_lang > 0:
        ws_total.cell(row=9, column=8, value=sv_lang)
    ws_total.cell(row=9, column=9, value=len(df_lang))

    # Ghi đè công thức GPA đạt cục bộ thay vì liên kết file ngoài [1]
    tb_sheet_name = 'TB CUA CAC NGANH' if 'TB CUA CAC NGANH' in wb.sheetnames else 'TB CỦA CÁC NGÀNH'
    ws_total.cell(row=7, column=11, value=f"='{tb_sheet_name}'!G16")
    ws_total.cell(row=8, column=11, value=f"='{tb_sheet_name}'!H16")
    ws_total.cell(row=9, column=11, value=f"='{tb_sheet_name}'!I16")

    # Điền rating table cho TONG HOP (bắt đầu từ row 15)
    fill_rating_table(ws_total, metrics_all, start_row=15)
    
    # 8. Ghi dữ liệu cho các sheet Ngành lẻ
    if 'CNTT' in wb.sheetnames:
        fill_rating_table(wb['CNTT'], metrics_cntt, start_row=8)
    if 'KINH TE' in wb.sheetnames:
        fill_rating_table(wb['KINH TE'], metrics_biz, start_row=8)
    if 'NGON NGU' in wb.sheetnames:
        fill_rating_table(wb['NGON NGU'], metrics_lang, start_row=8)

    # 9. Ghi dữ liệu cho sheet 'TB CỦA CÁC NGÀNH' hoặc 'TB CUA CAC NGANH'
    tb_sheet_name = 'TB CUA CAC NGANH' if 'TB CUA CAC NGANH' in wb.sheetnames else 'TB CỦA CÁC NGÀNH'
    if tb_sheet_name in wb.sheetnames:
        ws_tb = wb[tb_sheet_name]
        
        # Đảm bảo có cột Ngôn ngữ ở Col I
        ws_tb.cell(row=5, column=9, value="Ngôn ngữ")
        ws_tb.cell(row=5, column=9).font = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
        ws_tb.cell(row=5, column=9).fill = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
        ws_tb.cell(row=5, column=9).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for q in range(8):
            r = 6 + q
            c_f = ws_tb.cell(row=r, column=6, value=round(metrics_all[q]["avg"], 2))
            c_f.number_format = '0.00'
            c_g = ws_tb.cell(row=r, column=7, value=round(metrics_cntt[q]["avg"], 2))
            c_g.number_format = '0.00'
            c_h = ws_tb.cell(row=r, column=8, value=round(metrics_biz[q]["avg"], 2))
            c_h.number_format = '0.00'
            c_i = ws_tb.cell(row=r, column=9, value=round(metrics_lang[q]["avg"], 2))
            c_i.number_format = '0.00'

        # Điền trung bình cộng cuối bảng (Row 14: C1-C8, Row 15: C1-C4, Row 16: C1-C7)
        def fill_tb_summary_row(row_idx, q_indices):
            scores_all = [metrics_all[qi]["avg"] for qi in q_indices]
            avg_all = sum(scores_all) / len(scores_all) if scores_all else 0
            c_f = ws_tb.cell(row=row_idx, column=6, value=round(avg_all, 2))
            c_f.number_format = '0.00'
            
            scores_cntt = [metrics_cntt[qi]["avg"] for qi in q_indices]
            avg_cntt = sum(scores_cntt) / len(scores_cntt) if scores_cntt else 0
            c_g = ws_tb.cell(row=row_idx, column=7, value=round(avg_cntt, 2))
            c_g.number_format = '0.00'
            
            scores_biz = [metrics_biz[qi]["avg"] for qi in q_indices]
            avg_biz = sum(scores_biz) / len(scores_biz) if scores_biz else 0
            c_h = ws_tb.cell(row=row_idx, column=8, value=round(avg_biz, 2))
            c_h.number_format = '0.00'
            
            scores_lang = [metrics_lang[qi]["avg"] for qi in q_indices]
            avg_lang = sum(scores_lang) / len(scores_lang) if scores_lang else 0
            c_i = ws_tb.cell(row=row_idx, column=9, value=round(avg_lang, 2))
            c_i.number_format = '0.00'

        fill_tb_summary_row(14, range(8))
        fill_tb_summary_row(15, range(4))
        fill_tb_summary_row(16, range(7))

    # 10. Ghi dữ liệu cho sheet 'TY LE PHAN HOI'
    if 'TY LE PHAN HOI' in wb.sheetnames:
        ws_rate = wb['TY LE PHAN HOI']
        if sv_cntt > 0:
            ws_rate.cell(row=9, column=15, value=sv_cntt)
        ws_rate.cell(row=9, column=16, value=len(df_cntt))
        
        if sv_lang > 0:
            ws_rate.cell(row=10, column=15, value=sv_lang)
        ws_rate.cell(row=10, column=16, value=len(df_lang))
        
        if sv_biz > 0:
            ws_rate.cell(row=11, column=15, value=sv_biz)
        ws_rate.cell(row=11, column=16, value=len(df_biz))

    # 11. Clear và điền các ý kiến phản hồi thô vào các sheet 'Gop y Q*'
    q_to_sheet = {
        "Q1": "Gop y Q1",
        "Q2": "Gop y Q2",
        "Q3": "Gop y Q3",
        "Q4": "Gop y Q4",
        "Q5": "Gop y Q5"
    }
    
    from src.pipeline.data_loader import get_all_comments
    comments_all = get_all_comments(survey_df)
    
    for q_key, sname in q_to_sheet.items():
        if sname in wb.sheetnames:
            ws_cmt = wb[sname]
            if ws_cmt.max_row >= 3:
                ws_cmt.delete_rows(3, ws_cmt.max_row - 2)
            
            q_cmts = comments_all[comments_all["Question"] == q_key]
            
            # Tự động phát hiện số cột cần kẻ viền từ dòng header 2
            border_cols = 3
            for col_idx in range(1, 15):
                if ws_cmt.cell(row=2, column=col_idx).value is not None:
                    border_cols = col_idx

            stt = 1
            for _, row in q_cmts.iterrows():
                val_comment = row["Comment"]
                content = str(val_comment).strip() if pd.notna(val_comment) else ""
                if content:
                    r = 2 + stt
                    ws_cmt.cell(row=r, column=1, value=stt).alignment = Alignment(horizontal="center", vertical="top")
                    ws_cmt.cell(row=r, column=1).font = Font(name="Times New Roman", size=10)
                    
                    ws_cmt.cell(row=r, column=2, value=content).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    ws_cmt.cell(row=r, column=2).font = Font(name="Times New Roman", size=10)
                    
                    thin_border = Border(
                        left=Side(style='thin', color='BDBDBD'),
                        right=Side(style='thin', color='BDBDBD'),
                        top=Side(style='thin', color='BDBDBD'),
                        bottom=Side(style='thin', color='BDBDBD')
                    )
                    
                    # Áp border cho toàn bộ các cột của dòng dữ liệu góp ý
                    for col_idx in range(1, border_cols + 1):
                        ws_cmt.cell(row=r, column=col_idx).border = thin_border
                    
                    stt += 1

    # 12. Nếu có kết quả AI Grouping, tạo một sheet mới tên 'GOM NHÓM Ý KIẾN (AI)'
    if ai_group_result and "groups" in ai_group_result:
        if 'GOM NHÓM Ý KIẾN (AI)' in wb.sheetnames:
            wb.remove(wb['GOM NHÓM Ý KIẾN (AI)'])
            
        ws_ai = wb.create_sheet('GOM NHÓM Ý KIẾN (AI)')
        ws_ai.column_dimensions['A'].width = 6
        ws_ai.column_dimensions['B'].width = 25
        ws_ai.column_dimensions['C'].width = 14
        ws_ai.column_dimensions['D'].width = 14
        ws_ai.column_dimensions['E'].width = 10
        ws_ai.column_dimensions['F'].width = 50
        ws_ai.column_dimensions['G'].width = 50
        
        ws_ai.merge_cells('A1:G1')
        ws_ai.cell(row=1, column=1, value="GOM NHÓM & TÓM TẮT Ý KIẾN ĐÓNG GÓP (PHÂN TÍCH BẰNG AI)").font = Font(name="Times New Roman", bold=True, size=13, color="FFFFFF")
        ws_ai.cell(row=1, column=1).fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
        ws_ai.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_ai.row_dimensions[1].height = 32
        
        ws_ai.merge_cells('A2:G2')
        ws_ai.cell(row=2, column=1, value=f"{len(ai_group_result['groups'])} nhóm từ {ai_group_result.get('total_feedbacks', len(survey_df))} ý kiến  –  Cơ sở {coso}").font = Font(name="Times New Roman", italic=True, bold=True, size=11, color="FFFFFF")
        ws_ai.cell(row=2, column=1).fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        ws_ai.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_ai.row_dimensions[2].height = 22
        
        ws_ai.row_dimensions[3].height = 6
        
        headers = ["TT", "Tên nhóm", "Phạm vi câu", "Phân loại", "SL", "Tóm tắt chuyên gia", "Gợi ý phản hồi"]
        for col_idx, h in enumerate(headers):
            cell = ws_ai.cell(row=4, column=col_idx+1, value=h)
            cell.font = Font(name="Times New Roman", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin', color='BDBDBD'), right=Side(style='thin', color='BDBDBD'))
        ws_ai.row_dimensions[4].height = 28
        
        sentiment_vi = {"positive": "Tích cực", "negative": "Tiêu cực", "neutral": "Trung lập", "suggestion": "Đề xuất"}
        sentiment_color = {"positive": "34D399", "negative": "DC2626", "neutral": "9CA3AF", "suggestion": "7C3AED"}
        
        for gi, group in enumerate(ai_group_result["groups"]):
            r = 5 + gi
            ws_ai.row_dimensions[r].height = 48
            
            ws_ai.cell(row=r, column=1, value=gi+1).alignment = Alignment(horizontal="center", vertical="top")
            ws_ai.cell(row=r, column=1).font = Font(name="Times New Roman", bold=True, size=10, color="5B21B6")
            
            ws_ai.cell(row=r, column=2, value=group.get("name", "")).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_ai.cell(row=r, column=2).font = Font(name="Times New Roman", bold=True, size=10)
            
            ws_ai.cell(row=r, column=3, value=group.get("question_scope", "")).alignment = Alignment(horizontal="center", vertical="top")
            ws_ai.cell(row=r, column=3).font = Font(name="Times New Roman", size=10)
            
            sent = group.get("sentiment", "neutral")
            cell_sent = ws_ai.cell(row=r, column=4, value=sentiment_vi.get(sent, sent))
            cell_sent.alignment = Alignment(horizontal="center", vertical="top")
            cell_sent.font = Font(name="Times New Roman", bold=True, size=10, color=sentiment_color.get(sent, "000000"))
            
            ws_ai.cell(row=r, column=5, value=group.get("count", 0)).alignment = Alignment(horizontal="center", vertical="top")
            ws_ai.cell(row=r, column=5).font = Font(name="Times New Roman", bold=True, size=10, color="5B21B6")
            
            ws_ai.cell(row=r, column=6, value=group.get("summary", "")).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_ai.cell(row=r, column=6).font = Font(name="Times New Roman", size=10)
            
            ws_ai.cell(row=r, column=7, value=group.get("suggested_response", "")).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_ai.cell(row=r, column=7).font = Font(name="Times New Roman", size=10, color="047857")
            
            thin_border = Border(
                left=Side(style='thin', color='BDBDBD'),
                right=Side(style='thin', color='BDBDBD'),
                top=Side(style='thin', color='BDBDBD'),
                bottom=Side(style='thin', color='BDBDBD')
            )
            for c in range(1, 8):
                ws_ai.cell(row=r, column=c).border = thin_border
                
    # 13. Lưu lại file
    wb.save(output_path)
    return output_path
